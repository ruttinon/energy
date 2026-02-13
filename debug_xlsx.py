
import os
import sys
import openpyxl
from datetime import datetime

# Path to the specific Excel file for the active project
# Assuming current month is Jan 2026 based on previous prompts
XLSX_PATH = r"c:\Users\promb\Desktop\energylink3\projects\CPRAM-639ec8\data\2026_01.xlsx"

def inspect_excel():
    print(f"Inspecting File: {XLSX_PATH}")
    
    if not os.path.exists(XLSX_PATH):
        print("❌ File NOT FOUND")
        return

    try:
        wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)
        print("✅ Workbook Loaded Successfully")
        print(f"Sheets: {wb.sheetnames}")
        
        for sheet_name in ['Readings', 'Historical']:
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                print(f"\n--- Sheet: {sheet_name} ---")
                print(f"Max Row: {ws.max_row}")
                
                # Print headers
                headers = [c.value for c in ws[1]] if ws.max_row >= 1 else []
                print(f"Headers: {headers}")
                
                # Print last 5 rows
                print("Last 5 Rows:")
                rows = list(ws.iter_rows(values_only=True))
                start_idx = max(1, len(rows) - 5)
                for i in range(start_idx, len(rows)):
                    print(f"Row {i+1}: {rows[i]}")
            else:
                print(f"\n❌ Sheet '{sheet_name}' NOT FOUND")
                
    except Exception as e:
        print(f"❌ Error reading Excel: {e}")

if __name__ == "__main__":
    inspect_excel()
