from fastapi import APIRouter, UploadFile, File, HTTPException
from openpyxl import load_workbook
import io
import os
import json
import sys

# Determine ROOT to avoid circular import
if getattr(sys, 'frozen', False):
    ROOT = os.path.dirname(sys.executable)
else:
    # services/backend/api/import_tool.py -> .../energylink3
    ROOT = os.path.abspath(os.path.join(os.path.dirname(__file__), '..', '..', '..'))

PROJECTS_ROOT = os.path.join(ROOT, 'projects')

router = APIRouter()

@router.post("/projects/{project_id}/parse_import")
async def parse_import_excel(project_id: str, file: UploadFile = File(...)):
    if not file.filename.endswith('.xlsx'):
        raise HTTPException(status_code=400, detail="Invalid file format. Please upload .xlsx file.")

    content = await file.read()
    wb = load_workbook(filename=io.BytesIO(content), data_only=True)
    
    # 1. Load Existing Config to check for duplicates
    existing_map = {} # { "ConverterName": { set of addresses } }
    
    config_path = os.path.join(PROJECTS_ROOT, project_id, "ConfigDevice.json")
    if os.path.exists(config_path):
        try:
            with open(config_path, 'r', encoding='utf-8') as f:
                cfg = json.load(f)
                for conv in cfg.get('converters', []):
                    cname = conv.get('name')
                    c_addrs = set()
                    for d in conv.get('devices', []):
                        addr = d.get('modbus_slave') or d.get('address')
                        if addr:
                            c_addrs.add(int(addr))
                    existing_map[cname] = c_addrs
        except Exception:
            pass

    # 2. Parse Excel
    result = {
        "converters": [],
        "summary": {"total_devices": 0, "errors": 0}
    }

    # Expected Columns (Flexible matching)
    # Meter Name, Serial Number, Panel, Address, CT Ratio
    # We map loose headers to keys
    
    for sheet_name in wb.sheetnames:
        # User Rule: Sheet Name = Converter Name
        ws = wb[sheet_name]
        
        # Identify Header Row (assuming row 1 or 2, scan first few)
        header_row_idx = -1
        col_map = {} # { "address": col_idx, "name": col_idx, ... }
        
        # Scan first 50 rows for "Address" keyword
        print(f"[Import] Scanning sheet '{sheet_name}'...")
        for i, row in enumerate(ws.iter_rows(min_row=1, max_row=50, values_only=True), start=1):
            row_lower = [str(c).lower().strip() if c else "" for c in row]
            # Debug: print first few non-empty rows
            if any(row_lower):
                print(f"[Import] Row {i}: {row_lower}")
            
            # Check for Address column
            if "address" in row_lower:
                header_row_idx = i
                print(f"[Import] Found Header at Row {i}")
                # Map columns
                for idx, cell in enumerate(row_lower):
                    if "address" in cell: col_map["address"] = idx
                    elif "name" in cell and "meter" in cell: col_map["name"] = idx # Favor "Meter Name"
                    elif "name" in cell and "plate" not in cell: 
                        if "name" not in col_map: col_map["name"] = idx
                    elif "serial" in cell: col_map["serial"] = idx
                    elif "panel" in cell: col_map["panel"] = idx
                    elif "ratio" in cell: col_map["ct_ratio"] = idx
                    elif "model" in cell: col_map["model"] = idx
                break
        
        if header_row_idx == -1:
            print(f"[Import] No text 'Address' found in first 50 rows of '{sheet_name}'")
            continue
        
        if "address" not in col_map:
             print(f"[Import] Header row found but 'address' column mapping failed.")
             continue

        converter_obj = {
            "name": sheet_name,
            "devices": [],
            "id": None, # Will be assigned by frontend or match existing
            "exists": sheet_name in existing_map
        }
        
        # Addresses used in this sheet (to check internal duplicates)
        sheet_addresses = set()
        
        # Parse Rows
        for r_idx, row in enumerate(ws.iter_rows(min_row=header_row_idx+1, values_only=True), start=header_row_idx+1):
            # Check if row is empty
            if not any(row): continue
            
            # Extract Address
            addr_idx = col_map["address"]
            addr_raw = row[addr_idx] if addr_idx < len(row) else None
            
            if addr_raw is None: continue # valid row needs address? Or return error?
            
            # Extract Key Fields
            name_v = row[col_map["name"]] if "name" in col_map and col_map["name"] < len(row) else f"Meter_{addr_raw}"
            serial_v = row[col_map["serial"]] if "serial" in col_map and col_map["serial"] < len(row) else ""
            panel_v = row[col_map["panel"]] if "panel" in col_map and col_map["panel"] < len(row) else ""
            ct_v = row[col_map["ct_ratio"]] if "ct_ratio" in col_map and col_map["ct_ratio"] < len(row) else ""
            model_v = row[col_map["model"]] if "model" in col_map and col_map["model"] < len(row) else ""
            
            dev_obj = {
                "name": name_v,
                "serial_number": serial_v,
                "panel": panel_v,
                "address": addr_raw,
                "ct_ratio": ct_v,
                "model": model_v,
                "errors": [],
                "row": r_idx
            }
            
            # Validation
            try:
                addr_int = int(addr_raw)
                dev_obj["address_int"] = addr_int
                
                # Check duplicate in Sheet
                if addr_int in sheet_addresses:
                    dev_obj["errors"].append("Duplicate Address in File")
                else:
                    sheet_addresses.add(addr_int)
                
                # Check duplicate in Existing System
                if sheet_name in existing_map and addr_int in existing_map[sheet_name]:
                     # WARNING: User said "Not affect existing". Does that mean we skip? or Error?
                     # "ไม่กระทบ Device ที่มีอยู่เดิม" -> Don't overwrite. But effectively "Duplicate ID" if we try to add.
                     dev_obj["errors"].append("Address already exists in system")
                     
            except:
                dev_obj["errors"].append("Invalid Address (Must be integer)")

            if dev_obj["errors"]:
                result["summary"]["errors"] += 1
                
            converter_obj["devices"].append(dev_obj)
            result["summary"]["total_devices"] += 1

        if converter_obj["devices"]:
            result["converters"].append(converter_obj)
            
    return result
