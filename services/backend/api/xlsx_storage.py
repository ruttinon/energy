import os
from datetime import datetime, timedelta
from openpyxl import Workbook, load_workbook
import zipfile, time, uuid, os
import threading

_WB_LOCK = threading.RLock()
_LAST_ARCHIVE_TIME = {}

print("[XLSX_STORAGE] module loaded")

ROOT = os.path.abspath(os.path.join(os.path.dirname(__file__), '..', '..', '..'))
PROJECTS_ROOT = os.path.join(ROOT, 'projects')

def get_project_dir(project_id: str):
    return os.path.join(PROJECTS_ROOT, project_id)

def get_data_dir(project_id: str):
    """projects/CPRAM-639ec8/data/"""
    data_dir = os.path.join(PROJECTS_ROOT, project_id, 'data')
    os.makedirs(data_dir, exist_ok=True)
    return data_dir

def get_monthly_file(project_id: str, year: str, month: str):
    """projects/CPRAM-639ec8/data/2025_01.xlsx"""
    data_dir = get_data_dir(project_id)
    return os.path.join(data_dir, f"{year}_{month}.xlsx")

READINGS_HEADERS = ['date', 'time', 'device_id', 'device_name', 'key', 'value', 'unit']
BILLING_HEADERS = ['date', 'device_id', 'device_name', 'meter_start', 'meter_end', 'energy_used', 'price_per_unit', 'cost', 'last_update']
BILLING_WEEK_HEADERS = ['year', 'week', 'device_id', 'device_name', 'energy_used', 'cost']
BILLING_MONTH_HEADERS = ['month', 'device_id', 'device_name', 'energy_used', 'cost']
BILLING_YEAR_HEADERS = ['year', 'device_id', 'device_name', 'energy_used', 'cost']
HISTORICAL_HEADERS = ['timestamp', 'device_id', 'device_name', 'key', 'value', 'unit', 'interval_minutes']
SUMMARY_HEADERS = ['device_name', 'key', 'unit', 'readings_count', 'sum', 'avg', 'min', 'max']
ALERTS_HEADERS = ['timestamp', 'device_id', 'device_name', 'alert_type', 'message', 'resolved']

def ensure_sheets(filepath: str):
    """สร้างไฟล์ Excel พร้อม 4 sheets - WITH CORRUPTION RECOVERY"""
    try:
        if os.path.exists(filepath):
            try:
                wb = load_workbook(filepath)
                # ตรวจสอบว่ามี sheets ครบหรือไม่
                for name in ['Readings', 'Historical', 'Summary', 'Alerts', 'Billing', 'Billing_Weekly', 'Billing_Monthly', 'Billing_Yearly']:
                    if name not in wb.sheetnames:
                        ws = wb.create_sheet(name)
                        if name == 'Readings':
                            ws.append(READINGS_HEADERS)
                        elif name == 'Historical':
                            ws.append(HISTORICAL_HEADERS)
                        elif name == 'Summary':
                            ws.append(SUMMARY_HEADERS)
                        elif name == 'Alerts':
                            ws.append(ALERTS_HEADERS)
                        elif name == 'Billing':
                            ws.append(BILLING_HEADERS)
                        elif name == 'Billing_Weekly':
                            ws.append(BILLING_WEEK_HEADERS)
                        elif name == 'Billing_Monthly':
                            ws.append(BILLING_MONTH_HEADERS)
                        elif name == 'Billing_Yearly':
                            ws.append(BILLING_YEAR_HEADERS)
                wb.save(filepath)
                wb.close()
                return
            except Exception as e:
                # ไฟล์เสีย - ลบและสร้างใหม่
                print(f"[XLSX] Corrupted file detected, recreating: {filepath} ({e})")
                try:
                    os.remove(filepath)
                except:
                    pass
        
        # สร้างไฟล์ใหม่
        wb = Workbook()
        wb.remove(wb.active)
        
        ws = wb.create_sheet('Readings')
        ws.append(READINGS_HEADERS)
        
        ws = wb.create_sheet('Historical')
        ws.append(HISTORICAL_HEADERS)
        
        ws = wb.create_sheet('Summary')
        ws.append(SUMMARY_HEADERS)
        
        ws = wb.create_sheet('Alerts')
        ws.append(ALERTS_HEADERS)

        ws = wb.create_sheet('Billing')
        ws.append(BILLING_HEADERS)
        ws = wb.create_sheet('Billing_Weekly')
        ws.append(BILLING_WEEK_HEADERS)
        ws = wb.create_sheet('Billing_Monthly')
        ws.append(BILLING_MONTH_HEADERS)
        ws = wb.create_sheet('Billing_Yearly')
        ws.append(BILLING_YEAR_HEADERS)
        
        _safe_save(filepath, wb)
        print(f"[XLSX] Created new file: {filepath}")
        
    except Exception as e:
        print(f"[ERROR] ensure_sheets: {e}")

def _safe_load_workbook(filepath: str):
    """Load workbook with retries and auto-recreate when invalid zip."""
    for attempt in range(4):
        try:
            if not os.path.exists(filepath):
                ensure_sheets(filepath)
            # Validate zip structure
            if not zipfile.is_zipfile(filepath):
                print(f"[XLSX] Invalid zip structure, recreating: {filepath}")
                ensure_sheets(filepath)
                time.sleep(0.2)
            wb = load_workbook(filepath)
            return wb
        except Exception as e:
            if attempt == 3:
                # final fallback: recreate and try once
                ensure_sheets(filepath)
                try:
                    wb = load_workbook(filepath)
                    return wb
                except Exception:
                    raise
            time.sleep(0.2)
    # Should not reach here
    return load_workbook(filepath)

def _is_file_locked(filepath: str) -> bool:
    try:
        if not os.path.exists(filepath):
            return False
        f = open(filepath, 'rb+')
        try:
            f.flush()
        finally:
            f.close()
        return False
    except Exception:
        return True

def _cleanup_sibling_tmps(filepath: str):
    try:
        base = os.path.basename(filepath)
        dirp = os.path.dirname(filepath)
        prefix = base + '.'
        for name in os.listdir(dirp):
            if name.startswith(prefix) and name.endswith('.tmp'):
                try:
                    os.remove(os.path.join(dirp, name))
                except Exception:
                    pass
    except Exception:
        pass

def _safe_save(filepath: str, wb: Workbook) -> bool:
    try:
        if _is_file_locked(filepath):
            print(f"[XLSX] File locked, skip save: {filepath}")
            return False
        tmp = filepath + f".{uuid.uuid4().hex}.tmp"
        wb.save(tmp)
        wb.close()
        try:
            os.replace(tmp, filepath)
        except Exception:
            try:
                os.remove(tmp)
            except Exception:
                pass
            return False
        _cleanup_sibling_tmps(filepath)
        return True
    except Exception as e:
        try:
            wb.close()
        except Exception:
            pass
        print(f"[ERROR] _safe_save: {e}")
        return False

def append_row(filepath: str, sheet_name: str, row_data: dict, headers: list):
    """เพิ่มแถวข้อมูล"""
    try:
        with _WB_LOCK:
            ensure_sheets(filepath)
            wb = _safe_load_workbook(filepath)
            ws = wb[sheet_name]
            row = [row_data.get(h, '') for h in headers]
            ws.append(row)
            _safe_save(filepath, wb)
    except Exception as e:
        print(f"[ERROR] append_row {sheet_name}: {e}")

def upsert_row(filepath: str, sheet_name: str, match_keys: dict, row_data: dict, headers: list):
    try:
        with _WB_LOCK:
            ensure_sheets(filepath)
            wb = _safe_load_workbook(filepath)
            ws = wb[sheet_name]
            header_index = {h: i for i, h in enumerate(headers)}
            target_row_idx = None
            for ridx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                ok = True
                for mk, mv in match_keys.items():
                    idx = header_index.get(mk)
                    if idx is None:
                        ok = False
                        break
                    if str(row[idx]) != str(mv):
                        ok = False
                        break
                if ok:
                    target_row_idx = ridx
                    break
            values = [row_data.get(h, '') for h in headers]
            if target_row_idx:
                for c, v in enumerate(values, start=1):
                    ws.cell(row=target_row_idx, column=c, value=v)
            else:
                ws.append(values)
            _safe_save(filepath, wb)
    except Exception as e:
        print(f"[ERROR] upsert_row {sheet_name}: {e}")

def save_billing_daily(project_id: str, entry: dict):
    """Save daily billing entry to monthly Excel Billing sheet"""
    now = datetime.now()
    filepath = get_monthly_file(project_id, now.strftime('%Y'), now.strftime('%m'))
    row_data = {
        'date': entry.get('date') or now.strftime('%Y-%m-%d'),
        'device_id': entry.get('device_id'),
        'device_name': entry.get('device_name'),
        'meter_start': entry.get('meter_start'),
        'meter_end': entry.get('meter_end'),
        'energy_used': entry.get('energy_used'),
        'price_per_unit': entry.get('price_per_unit'),
        'cost': entry.get('cost'),
        'last_update': entry.get('last_update')
    }
    upsert_row(filepath, 'Billing', {'date': row_data['date'], 'device_id': row_data['device_id']}, row_data, BILLING_HEADERS)

def save_billing_weekly(project_id: str, year: str, week: str, device_id: str, device_name: str, energy: float, cost: float):
    filepath = get_monthly_file(project_id, year, datetime.now().strftime('%m'))
    row = {
        'year': year,
        'week': week,
        'device_id': device_id,
        'device_name': device_name,
        'energy_used': energy,
        'cost': cost,
    }
    upsert_row(filepath, 'Billing_Weekly', {'year': year, 'week': week, 'device_id': device_id}, row, BILLING_WEEK_HEADERS)

def save_billing_monthly(project_id: str, month: str, device_id: str, device_name: str, energy: float, cost: float):
    year = month.split('-')[0]
    filepath = get_monthly_file(project_id, year, month.split('-')[1])
    row = {
        'month': month,
        'device_id': device_id,
        'device_name': device_name,
        'energy_used': energy,
        'cost': cost,
    }
    upsert_row(filepath, 'Billing_Monthly', {'month': month, 'device_id': device_id}, row, BILLING_MONTH_HEADERS)

def save_billing_yearly(project_id: str, year: str, month: str, energy: float, cost: float):
    filepath = get_monthly_file(project_id, year, datetime.now().strftime('%m'))
    row = {
        'year': year,
        'device_id': '',
        'device_name': '',
        'energy_used': energy,
        'cost': cost,
    }
    upsert_row(filepath, 'Billing_Yearly', {'year': year}, row, BILLING_YEAR_HEADERS)

def save_reading(project_id: str, device_id: str, device_name: str, key: str, value, unit: str):
    """บันทึก realtime reading + auto-archive ทุก 15 นาที"""
    now = datetime.now()
    filepath = get_monthly_file(project_id, now.strftime('%Y'), now.strftime('%m'))
    
    row_data = {
        'date': now.strftime('%Y-%m-%d'),
        'time': now.strftime('%H:%M:%S'),
        'device_id': device_id,
        'device_name': device_name,
        'key': key,
        'value': value,
        'unit': unit
    }
    upsert_row(filepath, 'Readings', {'device_id': device_id, 'key': key}, row_data, READINGS_HEADERS)
    
    last_time = _LAST_ARCHIVE_TIME.get(project_id, now - timedelta(minutes=20))
    if (now - last_time).total_seconds() >= 15 * 60:
        archive_readings_to_historical(project_id)
        _LAST_ARCHIVE_TIME[project_id] = now

def save_readings_batch(project_id: str, readings_list: list):
    now = datetime.now()
    filepath = get_monthly_file(project_id, now.strftime('%Y'), now.strftime('%m'))
    try:
        with _WB_LOCK:
            ensure_sheets(filepath)
            wb = _safe_load_workbook(filepath)
            for r in readings_list:
                row_data = {
                    'date': now.strftime('%Y-%m-%d'),
                    'time': now.strftime('%H:%M:%S'),
                    'device_id': r.get('device_id'),
                    'device_name': r.get('device_name'),
                    'key': r.get('key'),
                    'value': r.get('value'),
                    'unit': r.get('unit', '')
                }
                upsert_row(filepath, 'Readings', {'device_id': row_data['device_id'], 'key': row_data['key']}, row_data, READINGS_HEADERS)
        last_time = _LAST_ARCHIVE_TIME.get(project_id, now - timedelta(minutes=20))
        if (now - last_time).total_seconds() >= 15 * 60:
            archive_readings_to_historical(project_id)
            _LAST_ARCHIVE_TIME[project_id] = now
            try:
                generate_monthly_summary(project_id, now.strftime('%Y'), now.strftime('%m'))
            except Exception as e:
                print(f"[WARN] summary generation failed: {e}")
    except Exception as e:
        print(f"[ERROR] save_readings_batch: {e}")

def archive_readings_to_historical(project_id: str):
    """ย้ายข้อมูลจาก Readings → Historical แล้วลบ Readings"""
    try:
        now = datetime.now()
        filepath = get_monthly_file(project_id, now.strftime('%Y'), now.strftime('%m'))
        
        if not os.path.exists(filepath):
            return
        
        with _WB_LOCK:
            wb = load_workbook(filepath)
            ws_readings = wb['Readings']
            ws_historical = wb['Historical']
            
            rows_to_archive = []
            for row in ws_readings.iter_rows(min_row=2, values_only=True):
                if row and row[0]:
                    rows_to_archive.append(row)
            
            for date_val, time_val, device_id, device_name, key, value, unit in rows_to_archive:
                try:
                    timestamp_str = f"{date_val}T{time_val}"
                    row_data = {
                        'timestamp': timestamp_str,
                        'device_id': device_id,
                        'device_name': device_name,
                        'key': key,
                        'value': value,
                        'unit': unit,
                        'interval_minutes': 15
                    }
                    row = [row_data.get(h, '') for h in HISTORICAL_HEADERS]
                    ws_historical.append(row)
                except Exception as e:
                    print(f"[ERROR] archive row: {e}")
            
            while ws_readings.max_row > 1:
                ws_readings.delete_rows(2)
            
            wb.save(filepath)
            wb.close()
            
            print(f"[XLSX] Archived {len(rows_to_archive)} rows: {project_id}")
    except Exception as e:
        print(f"[ERROR] archive_readings_to_historical: {e}")

def save_historical_reading(project_id: str, device_id: str, device_name: str, key: str, value, unit: str, timestamp: str = None, interval_minutes: int = 15):
    """บันทึก historical reading โดยตรง"""
    now = datetime.now()
    filepath = get_monthly_file(project_id, now.strftime('%Y'), now.strftime('%m'))
    
    row_data = {
        'timestamp': timestamp or now.isoformat(),
        'device_id': device_id,
        'device_name': device_name,
        'key': key,
        'value': value,
        'unit': unit,
        'interval_minutes': interval_minutes
    }
    append_row(filepath, 'Historical', row_data, HISTORICAL_HEADERS)

def save_alert(project_id: str, device_id: str, device_name: str, alert_type: str, message: str):
    """บันทึก alert"""
    now = datetime.now()
    filepath = get_monthly_file(project_id, now.strftime('%Y'), now.strftime('%m'))
    
    row_data = {
        'timestamp': now.isoformat(),
        'device_id': device_id,
        'device_name': device_name,
        'alert_type': alert_type,
        'message': message,
        'resolved': 'N'
    }
    append_row(filepath, 'Alerts', row_data, ALERTS_HEADERS)

def save_device_status(project_id: str, device_id: str, device_name: str, status: str, last_reading: str):
    """บันทึก device status - ไม่ต้องทำอะไร"""
    pass

def generate_monthly_summary(project_id: str, year: str, month: str):
    """สรุป Readings + Historical → Summary sheet"""
    filepath = get_monthly_file(project_id, year, month)
    if not os.path.exists(filepath):
        return
    
    try:
        with _WB_LOCK:
            wb = _safe_load_workbook(filepath)
            ws_readings = wb['Readings']
            ws_historical = wb['Historical']
            ws_summary = wb['Summary']
            
            summary_data = {}
            
            # อ่าน Readings
            for row in ws_readings.iter_rows(min_row=2, values_only=True):
                if not row or not row[0]:
                    continue
                date, time_val, device_id, device_name, key, value, unit = row
                try:
                    value = float(value)
                except:
                    continue
                key_name = f"{device_name}.{key}"
                if key_name not in summary_data:
                    summary_data[key_name] = {'device_name': device_name, 'key': key, 'unit': unit,
                                             'count': 0, 'sum': 0, 'min': float('inf'), 'max': float('-inf')}
                summary_data[key_name]['count'] += 1
                summary_data[key_name]['sum'] += value
                summary_data[key_name]['min'] = min(summary_data[key_name]['min'], value)
                summary_data[key_name]['max'] = max(summary_data[key_name]['max'], value)
            
            # อ่าน Historical
            for row in ws_historical.iter_rows(min_row=2, values_only=True):
                if not row or not row[0]:
                    continue
                timestamp, device_id, device_name, key, value, unit, interval = row
                try:
                    value = float(value)
                except:
                    continue
                key_name = f"{device_name}.{key}"
                if key_name not in summary_data:
                    summary_data[key_name] = {'device_name': device_name, 'key': key, 'unit': unit,
                                             'count': 0, 'sum': 0, 'min': float('inf'), 'max': float('-inf')}
                summary_data[key_name]['count'] += 1
                summary_data[key_name]['sum'] += value
                summary_data[key_name]['min'] = min(summary_data[key_name]['min'], value)
                summary_data[key_name]['max'] = max(summary_data[key_name]['max'], value)
            
            # ลบ Summary เก่า
            while ws_summary.max_row > 1:
                ws_summary.delete_rows(2)
            
            # เขียน Summary ใหม่
            for key, data in summary_data.items():
                avg = data['sum'] / data['count'] if data['count'] > 0 else 0
                row_data = {
                    'device_name': data['device_name'],
                    'key': data['key'],
                    'unit': data['unit'],
                    'readings_count': data['count'],
                    'sum': round(data['sum'], 2),
                    'avg': round(avg, 2),
                    'min': round(data['min'], 2),
                    'max': round(data['max'], 2)
                }
                ws_summary.append([row_data.get(h, '') for h in SUMMARY_HEADERS])
            
            _safe_save(filepath, wb)
    except Exception as e:
        print(f"[ERROR] generate_monthly_summary: {e}")

def auto_rotate_year(project_id: str):
    """เมื่อถึง 1 มกราคม ให้ archive ปีเก่า"""
    now = datetime.now()
    if now.month == 1 and now.day <= 2:
        last_year = str(int(now.strftime('%Y')) - 1)
        print(f"[XLSX] Year rotated: {project_id} ({last_year})")

def _parse_dt(s: str):
    """Parse datetime string - handles both ISO and space-separated formats"""
    try:
        if not s:
            return None
        if isinstance(s, datetime):
            return s
        s_str = str(s).strip()
        
        # Remove timezone info if present (e.g., 2025-12-09T16:59:59.000Z -> 2025-12-09T16:59:59)
        if s_str.endswith('Z'):
            s_str = s_str[:-1]
        if '.000' in s_str:
            s_str = s_str.split('.')[0]
        
        if 'T' in s_str:
            return datetime.strptime(s_str, "%Y-%m-%dT%H:%M:%S")
        else:
            return datetime.strptime(s_str, "%Y-%m-%d %H:%M:%S")
    except Exception:
        try:
            s_clean = str(s).replace('.000Z', '').replace('Z', '')
            return datetime.fromisoformat(s_clean)
        except:
            return None

def _month_iter(start_dt: datetime, end_dt: datetime):
    y = start_dt.year
    m = start_dt.month
    while y < end_dt.year or (y == end_dt.year and m <= end_dt.month):
        yield (str(y), f"{m:02d}")
        m += 1
        if m > 12:
            m = 1
            y += 1

def read_history_from_excel(project_id: str, device_id: str, key: str, start_ts: str, end_ts: str, device_name: str = None):
    """
    🔥 FIXED VERSION - รองรับการจับคู่ทั้ง device_id และ device_name + Timezone handling
    """
    try:
        print(f"[XLSX] Query: dev_id={device_id} key={key} name={device_name} start={start_ts} end={end_ts}")
        
        sdt = _parse_dt(start_ts)
        edt = _parse_dt(end_ts)
        
        if not sdt or not edt:
            print("[XLSX] Invalid date range, using default 7 days")
            edt = datetime.now()
            sdt = edt - timedelta(days=7)
        
        # 🔥 FIX: เพิ่มเวลาให้ครอบคลุมทั้งวัน
        # ถ้า end time เป็นเที่ยงคืน ให้เป็นสิ้นวันแทน
        if edt.hour == 16 and edt.minute == 59:  # Frontend ส่งมาเป็น 16:59:59 UTC
            edt = edt.replace(hour=23, minute=59, second=59)
        
        print(f"[XLSX] Date range after parse: {sdt} to {edt}")
        
        out = []
        
        with _WB_LOCK:
            for y, m in _month_iter(sdt, edt):
                fp = get_monthly_file(project_id, y, m)
                print(f"[XLSX] Checking file: {fp}")
                
                if not os.path.exists(fp):
                    print(f"[XLSX] File not found: {fp}")
                    continue
                
                try:
                    wb = load_workbook(fp, data_only=True)
                except Exception as e:
                    print(f"[XLSX] Cannot load file (corrupted): {fp} - {e}")
                    # ลองสร้างใหม่
                    ensure_sheets(fp)
                    continue
                
                sheets_to_check = []
                if 'Historical' in wb.sheetnames:
                    sheets_to_check.append(('Historical', 'hist'))
                if 'Readings' in wb.sheetnames:
                    sheets_to_check.append(('Readings', 'real'))
                
                for sheet_name, sheet_type in sheets_to_check:
                    ws = wb[sheet_name]
                    for row in ws.iter_rows(min_row=2, values_only=True):
                        try:
                            if sheet_type == 'hist':
                                # Historical: timestamp, device_id, device_name, key, value, unit, interval
                                ts, did, dname, rkey, val, unit, interval = row
                                dt = _parse_dt(ts)
                            else:
                                # Readings: date, time, device_id, device_name, key, value, unit
                                date_val, time_val, did, dname, rkey, val, unit = row
                                dt = _parse_dt(f"{date_val}T{time_val}")

                            # 🔥 FIX: จับคู่ทั้ง ID และ Name + DEBUG
                            is_dev_match = False
                            
                            # ลองจับคู่ device_id
                            if str(did) == str(device_id):
                                is_dev_match = True
                                print(f"[XLSX] ✓ Match by device_id: {did} == {device_id}")
                            
                            # ถ้าไม่ match และมี device_name ให้ลองจับคู่
                            if not is_dev_match and device_name:
                                # ลองจับ device_id column เป็น name
                                if str(did).strip() == str(device_name).strip():
                                    is_dev_match = True
                                    print(f"[XLSX] ✓ Match by device_id->name: {did} == {device_name}")
                                # ลองจับ device_name column
                                if str(dname).strip() == str(device_name).strip():
                                    is_dev_match = True
                                    print(f"[XLSX] ✓ Match by device_name: {dname} == {device_name}")
                            
                            # ถ้ายังไม่ match ให้ลองจับ device_name column เป็น ID
                            if not is_dev_match and dname:
                                if str(dname) == str(device_id):
                                    is_dev_match = True
                                    print(f"[XLSX] ✓ Match by device_name->id: {dname} == {device_id}")
                            
                            if not is_dev_match:
                                # DEBUG: แสดงว่าทำไมไม่ match
                                if len(out) < 3:  # แสดงแค่ 3 rows แรกเพื่อไม่ให้ spam
                                    print(f"[XLSX] ✗ Skip row: did={did}, dname={dname} (looking for id={device_id}, name={device_name})")
                                continue
                            
                            # ต้อง match key ด้วย
                            key_match = str(rkey).strip() == str(key).strip()
                            if not key_match:
                                if len(out) < 3:
                                    print(f"[XLSX] ✗ Key mismatch: '{rkey}' != '{key}'")
                                continue
                            
                            print(f"[XLSX] ✓ Found match: {dname}/{rkey} = {val} @ {dt}")
                            
                            if not dt:
                                continue
                            
                            if dt < sdt or dt > edt:
                                continue
                                
                            v = float(val)
                            out.append({"timestamp": dt.strftime("%Y-%m-%d %H:%M:%S"), "value": v})
                            
                        except Exception as e:
                            continue
                
                wb.close()
        
        # Deduplicate
        seen = set()
        unique_out = []
        for item in out:
            if item['timestamp'] not in seen:
                seen.add(item['timestamp'])
                unique_out.append(item)
                
        unique_out.sort(key=lambda x: x.get("timestamp"))
        print(f"[XLSX] ✅ Found {len(unique_out)} records")
        return unique_out
        
    except Exception as e:
        print(f"[ERROR] read_history_from_excel: {e}")
        import traceback
        traceback.print_exc()
def read_alerts_from_excel(project_id: str, year: str = None, month: str = None):
    """
    Read alerts from the 'Alerts' sheet of the monthly Excel file.
    If year/month not provided, defaults to current month.
    """
    try:
        if not year or not month:
            now = datetime.now()
            year = now.strftime('%Y')
            month = now.strftime('%m')
            
        fp = get_monthly_file(project_id, year, month)
        if not os.path.exists(fp):
            return []
            
        alerts = []
        with _WB_LOCK:
            wb = load_workbook(fp, data_only=True)
            if 'Alerts' not in wb.sheetnames:
                wb.close()
                return []
                
            ws = wb['Alerts']
            # ALERTS_HEADERS = ['timestamp', 'device_id', 'device_name', 'alert_type', 'message', 'resolved']
            
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row or not row[0]:
                    continue
                    
                # Unpack row safely
                try:
                    timestamp, device_id, device_name, alert_type, message, resolved = row[:6]
                except ValueError:
                    continue # Skip malformed rows
                    
                alerts.append({
                    "timestamp": timestamp,
                    "device_id": device_id,
                    "device_name": device_name,
                    "severity": alert_type, # Mapping alert_type to severity
                    "message": message,
                    "resolved": resolved
                })
            wb.close()
            
        # Sort by timestamp descending
        alerts.sort(key=lambda x: x.get("timestamp") or "", reverse=True)
        return []
        
    except Exception as e:
        print(f"[ERROR] read_alerts_from_excel: {e}")
        return []

def get_billing_summary_from_excel(project_id: str, year: str = None, month: str = None):
    """Return summary from Billing sheet: today totals and month totals"""
    try:
        now = datetime.now()
        y = year or now.strftime('%Y')
        m = month or now.strftime('%m')
        today_str = now.strftime('%Y-%m-%d')
        fp = get_monthly_file(project_id, y, m)
        ensure_sheets(fp)
        today_units = 0.0
        today_money = 0.0
        month_units = 0.0
        month_money = 0.0
        with _WB_LOCK:
            wb = load_workbook(fp, data_only=True)
            if 'Billing' in wb.sheetnames:
                ws = wb['Billing']
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if not row:
                        continue
                    date, device_id, device_name, meter_start, meter_end, energy_used, price_per_unit, cost, last_update = (list(row) + [None]*9)[:9]
                    try:
                        e = float(energy_used or 0)
                        c = float(cost or 0)
                    except Exception:
                        e = 0.0
                        c = 0.0
                    month_units += e
                    month_money += c
                    if str(date) == today_str:
                        today_units += e
                        today_money += c
            wb.close()
        result = {
            'today_units': round(today_units, 3),
            'today_money': round(today_money, 2),
            'month_units': round(month_units, 3),
            'month_money': round(month_money, 2),
        }
        # Fallback to DB aggregates if Excel is empty
        if result['today_units'] == 0.0 and result['month_units'] == 0.0 and result['month_money'] == 0.0:
            try:
                from services.backend.api.billing.billing_service import get_dashboard_summary
                dbs = get_dashboard_summary(project_id)
                return {
                    'today_units': dbs.get('today_units', 0.0),
                    'today_money': dbs.get('today_money', 0.0),
                    'month_units': dbs.get('month_units', 0.0),
                    'month_money': dbs.get('month_money', 0.0),
                }
            except Exception:
                pass
        return result
    except Exception as e:
        print(f"[ERROR] get_billing_summary_from_excel: {e}")
        return {'today_units': 0.0, 'today_money': 0.0, 'month_units': 0.0, 'month_money': 0.0}

def get_daily_series_from_excel(project_id: str, month: str):
    try:
        if not month:
            now = datetime.now()
            month = now.strftime('%Y-%m')
        y = month.split('-')[0]
        m = month.split('-')[1]
        fp = get_monthly_file(project_id, y, m)
        ensure_sheets(fp)
        agg = {}
        with _WB_LOCK:
            wb = load_workbook(fp, data_only=True)
            if 'Billing' in wb.sheetnames:
                ws = wb['Billing']
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if not row:
                        continue
                    date, device_id, device_name, meter_start, meter_end, energy_used, price_per_unit, cost, last_update = (list(row) + [None]*9)[:9]
                    ds = str(date or '')
                    if not ds.startswith(month):
                        continue
                    try:
                        dnum = int(ds.split('-')[2])
                    except Exception:
                        continue
                    e = float(energy_used or 0)
                    c = float(cost or 0)
                    if dnum not in agg:
                        agg[dnum] = {'day': dnum, 'total_energy': 0.0, 'total_cost': 0.0}
                    agg[dnum]['total_energy'] += e
                    agg[dnum]['total_cost'] += c
            wb.close()
        series = [
            {'day': k, 'total_energy': round(v['total_energy'], 3), 'total_cost': round(v['total_cost'], 2)}
            for k, v in sorted(agg.items())
        ]
        # Fallback to DB series if Excel empty
        if not series:
            try:
                from services.backend.api.billing.billing_storage import get_daily_series
                dbs = get_daily_series(project_id, month)
                return dbs
            except Exception:
                pass
        return series
    except Exception as e:
        print(f"[ERROR] get_daily_series_from_excel: {e}")
        return []
