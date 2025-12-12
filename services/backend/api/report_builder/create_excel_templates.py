import os
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side # type: ignore

BASE_DIR = Path(__file__).resolve().parents[2]
OUT_DIR = BASE_DIR / "data" / "report_templates" / "excel_templates"
OUT_DIR.mkdir(parents=True, exist_ok=True)

title_font = Font(size=16, bold=True, color="FFFFFF")
header_fill = PatternFill("solid", fgColor="333333")
title_fill = PatternFill("solid", fgColor="C9184A")
thin = Side(style="thin", color="666666")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
center = Alignment(horizontal="center")
right = Alignment(horizontal="right")
left = Alignment(horizontal="left")

def set_cols(ws):
    ws.column_dimensions['A'].width = 28
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 18

def style_header(ws, row, titles):
    for i, t in enumerate(titles, start=1):
        cell = ws.cell(row=row, column=i, value=t)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.border = border
        cell.alignment = center

def add_title(ws, text):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)
    c = ws.cell(row=1, column=1, value=text)
    c.font = title_font
    c.fill = title_fill
    c.alignment = center

def daily_all_template():
    wb = Workbook()
    ws = wb.active
    ws.title = "Daily Overview"
    set_cols(ws)
    add_title(ws, "รายงานรายวัน {{date}} (โหมด: {{period_type}})")
    ws.append(["สรุปทั้งหมด", "", "", ""])
    ws.append(["จำนวนอุปกรณ์", "{{devices.length}}", "พลังงานรวม (kWh)", "{{total_used}}"])
    ws.append(["ยอดเงินรวม (บาท)", "{{total_money}}", "ราคาต่อหน่วย", "{{price_per_unit}}"])
    style_header(ws, 5, ["อุปกรณ์", "พลังงาน (kWh)", "ค่าไฟ (บาท)", "มิเตอร์ปัจจุบัน"])
    row = 6
    for i in range(1, 21):
        ws.append([f"{{{{devices.{i-1}.device_name}}}}", f"{{{{devices.{i-1}.used}}}}", f"{{{{devices.{i-1}.money}}}}", f"{{{{devices.{i-1}.meter_now}}}}"])
        for col in range(1,5):
            cell = ws.cell(row=row, column=col)
            cell.border = border
            cell.alignment = left if col==1 else right
        row += 1
    path = OUT_DIR / "daily_all.xlsx"
    wb.save(path)

def monthly_all_template():
    wb = Workbook()
    ws = wb.active
    ws.title = "Monthly Overview"
    set_cols(ws)
    add_title(ws, "รายงานรายเดือน {{date}}")
    ws.append(["สรุปทั้งเดือน", "", "", ""])
    ws.append(["พลังงานรวมเดือน (kWh)", "{{month_total_units}}", "ยอดเงินรวมเดือน (บาท)", "{{month_total_money}}"])
    style_header(ws, 4, ["Converter", "จำนวนอุปกรณ์", "พลังงานรวม", "ค่าไฟรวม"])
    row = 5
    for i in range(1, 21):
        ws.append([f"{{{{converters.{i-1}.name}}}}", f"{{{{converters.{i-1}.devices.length}}}}", f"{{{{converters.{i-1}.used}}}}", f"{{{{converters.{i-1}.money}}}}"])
        for col in range(1,5):
            cell = ws.cell(row=row, column=col)
            cell.border = border
            cell.alignment = left if col==1 else right
        row += 1
    path = OUT_DIR / "monthly_all.xlsx"
    wb.save(path)

def yearly_all_template():
    wb = Workbook()
    ws = wb.active
    ws.title = "Yearly Overview"
    set_cols(ws)
    add_title(ws, "รายงานรายปี {{date}}")
    ws.append(["พลังงานรวมปี (kWh)", "{{month_total_units}}", "ยอดเงินรวมปี (บาท)", "{{month_total_money}}"])
    style_header(ws, 3, ["อุปกรณ์", "พลังงานรวม", "ค่าไฟรวม"])
    path = OUT_DIR / "yearly_all.xlsx"
    wb.save(path)

def converter_template():
    wb = Workbook()
    ws = wb.active
    ws.title = "Converter"
    set_cols(ws)
    add_title(ws, "รายงาน Converter {{converter_name}} — วันที่ {{date}}")
    ws.append(["จำนวนอุปกรณ์", "{{devices.length}}", "พลังงานรวม (kWh)", "{{total_used}}"])
    ws.append(["ค่าไฟรวม (บาท)", "{{total_money}}", "ราคาต่อหน่วย", "{{price_per_unit}}"])
    style_header(ws, 4, ["อุปกรณ์", "พลังงาน (kWh)", "ค่าไฟ (บาท)", "มิเตอร์ปัจจุบัน"])
    row = 5
    for i in range(1, 21):
        ws.append([f"{{{{devices.{i-1}.device_name}}}}", f"{{{{devices.{i-1}.used}}}}", f"{{{{devices.{i-1}.money}}}}", f"{{{{devices.{i-1}.meter_now}}}}"])
        for col in range(1,5):
            cell = ws.cell(row=row, column=col)
            cell.border = border
            cell.alignment = left if col==1 else right
        row += 1
    path = OUT_DIR / "converter_group.xlsx"
    wb.save(path)

def device_single_template():
    wb = Workbook()
    ws = wb.active
    ws.title = "Device"
    set_cols(ws)
    add_title(ws, "รายงานอุปกรณ์เดี่ยว {{device_name}} — วันที่ {{date}}")
    ws.append(["Converter", "{{convertor_name}}", "รหัสอุปกรณ์", "{{device_id}}"])
    ws.append(["มิเตอร์เริ่มต้น", "{{meter_prev}}", "มิเตอร์ปัจจุบัน", "{{meter_now}}"])
    ws.append(["พลังงานที่ใช้ (kWh)", "{{used}}", "ราคาต่อหน่วย", "{{price_per_unit}}"])
    ws.append(["ค่าไฟรายวัน (บาท)", "{{money}}", "รวมค่าไฟ (บาท)", "{{total_money}}"])
    path = OUT_DIR / "device_single.xlsx"
    wb.save(path)

def main():
    daily_all_template()
    monthly_all_template()
    yearly_all_template()
    converter_template()
    device_single_template()
    print("✅ Created templates in", OUT_DIR)

if __name__ == "__main__":
    main()
