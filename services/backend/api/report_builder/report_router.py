
from fastapi import APIRouter, File, UploadFile, HTTPException, Query, Body, Response
from fastapi.responses import JSONResponse, FileResponse, StreamingResponse
import os
import json
import shutil
import io
import csv
import pandas as pd
from datetime import datetime, timedelta
from pathlib import Path
# from services.backend.api import xlsx_storage as storage # REMOVED
import openpyxl
from openpyxl.utils import get_column_letter

router = APIRouter()

# ...

# REPLACEMENT PATH UTILS
def get_project_dir_local(project_id: str):
    if getattr(sys, 'frozen', False):
        root = os.path.dirname(sys.executable)
    else:
        root = os.path.abspath(os.path.join(os.path.dirname(__file__), '..', '..', '..', '..'))
    return os.path.join(root, 'projects', project_id)

def get_data_dir_local(project_id: str):
    return os.path.join(get_project_dir_local(project_id), 'data')

def get_project_report_dir(project_id: str):
    d = os.path.join(get_data_dir_local(project_id), "reports")
    os.makedirs(d, exist_ok=True)
    return d

def get_templates_dir(project_id: str = None):
    # If project_id is provided, try to find project specific templates
    if project_id:
        d = os.path.join(get_project_dir_local(project_id), "templates report")
        os.makedirs(d, exist_ok=True)
        return d
    return TEMPLATES_DIR

def load_json(path, default=None):
    if default is None: default = {}
    if not os.path.exists(path): return default
    try:
        with open(path, "r", encoding="utf-8") as f:
            return json.load(f)
    except:
        return default

# ==========================================================
# TEMPLATE MANAGEMENT
# ==========================================================

@router.get("/excel/templates")
def list_excel_templates(project_id: str = Query(None)):
    templates = []
    
    # 1. System Templates
    sys_dir = TEMPLATES_DIR
    if os.path.exists(sys_dir):
        for f in os.listdir(sys_dir):
            if f.endswith(".xlsx") and not f.startswith("~$"):
                path = os.path.join(sys_dir, f)
                stat = os.stat(path)
                templates.append({
                    "id": f, 
                    "filename": f,
                    "size": stat.st_size,
                    "modified": stat.st_mtime,
                    "scope": "system"
                })

    # 2. Project Templates (override system if same name? or distinct?)
    # Let's say we list them all.
    if project_id:
        prj_dir = get_templates_dir(project_id)
        if os.path.exists(prj_dir) and prj_dir != sys_dir:
             for f in os.listdir(prj_dir):
                if f.endswith(".xlsx") and not f.startswith("~$"):
                    path = os.path.join(prj_dir, f)
                    stat = os.stat(path)
                    # Check if exists in system
                    existing = next((t for t in templates if t['filename'] == f), None)
                    if existing:
                        existing['scope'] = 'project_override' # Mark as overriding
                        existing['modified'] = stat.st_mtime
                    else:
                        templates.append({
                            "id": f, 
                            "filename": f,
                            "size": stat.st_size,
                            "modified": stat.st_mtime,
                            "scope": "project"
                        })
    
    return {"templates": templates}

@router.post("/excel/template/upload")
def upload_excel_template(file: UploadFile = File(...), project_id: str = Query(None)):
    if not file.filename.endswith(".xlsx"):
        raise HTTPException(400, "Only .xlsx files allowed")
    
    target_dir = get_templates_dir(project_id)
    path = os.path.join(target_dir, file.filename)
    
    # If using system dir, maybe we restrict? For now allow overwrite.
    
    with open(path, "wb") as buffer:
        shutil.copyfileobj(file.file, buffer)
        
    return {"status": "ok", "filename": file.filename, "scope": "project" if project_id else "system"}

@router.delete("/excel/template/{filename}")
def delete_excel_template(filename: str, project_id: str = Query(None)):
    # Try project dir first
    if project_id:
        target_dir = get_templates_dir(project_id)
        path = os.path.join(target_dir, filename)
        if os.path.exists(path):
            os.remove(path)
            return {"status": "ok", "scope": "project"}
    
    # Then system (maybe disable deleting system templates via API? Safe to allow for now)
    path = os.path.join(TEMPLATES_DIR, filename)
    if os.path.exists(path):
        os.remove(path)
        return {"status": "ok", "scope": "system"}
    raise HTTPException(404, "Template not found")

@router.get("/excel/template/{filename}/content")
def get_template_content(filename: str, project_id: str = Query(None)):
    # Priority: Project > System
    path = None
    if project_id:
        p_dir = get_templates_dir(project_id)
        p_path = os.path.join(p_dir, filename)
        if os.path.exists(p_path):
            path = p_path
    
    if not path:
        path = os.path.join(TEMPLATES_DIR, filename)

    if not os.path.exists(path):
        raise HTTPException(404, "Template not found")
    
    try:
        print(f"[EXCEL] Loading template: {path}")
        wb = openpyxl.load_workbook(path, data_only=False) # Keep formulas
        ws = wb.active 
        print(f"[EXCEL] Active sheet: {ws.title}, max_row={ws.max_row}, max_col={ws.max_column}")
        
        data = []
        # Dynamic limits based on content
        # Reduce max limits to prevent frontend lag
        max_row = min(ws.max_row, 100) 
        max_col = min(ws.max_column, 26) 
        
        # Ensure minimal grid (smaller default)
        max_row = max(max_row, 20)
        max_col = max(max_col, 10)

        for row in ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col, values_only=True):
            r_data = []
            for cell in row:
                if cell is None:
                    r_data.append("")
                else:
                    r_data.append(str(cell))
            data.append(r_data)
            
        print(f"[EXCEL] Loaded {len(data)} rows")
        return {"status": "ok", "data": data, "rows": max_row, "cols": max_col}
    except Exception as e:
         print(f"[EXCEL] Error loading {filename}: {e}")
         return JSONResponse({"status": "error", "msg": str(e)}, status_code=500)

@router.post("/excel/template/{filename}/save")
def save_template_content(filename: str, payload: dict = Body(...), project_id: str = Query(None)):
    
    # Save ALWAYS to project dir if project_id exists (copy-on-write)
    if project_id:
        target_dir = get_templates_dir(project_id)
        path = os.path.join(target_dir, filename)
        
        # If not exists in project but exists in system, copy it first
        if not os.path.exists(path):
             sys_path = os.path.join(TEMPLATES_DIR, filename)
             if os.path.exists(sys_path):
                 shutil.copy2(sys_path, path)
             else:
                 # New file
                 pass 
    else:
        # System save
        path = os.path.join(TEMPLATES_DIR, filename)

    if not os.path.exists(path):
        # We can probably create new empty if missing?
        # raise HTTPException(404, "Template not found")
        # Let's verify if user intends to create new? The openpyxl load needs it.
        # Fallback create empty?
        wb = openpyxl.Workbook()
        wb.save(path)
    
    try:
        wb = openpyxl.load_workbook(path)
        ws = wb.active
        
        grid_data = payload.get("data", [])
        
        for r_idx, row_val in enumerate(grid_data):
            for c_idx, cell_val in enumerate(row_val):
                # 1-based index
                # Only update if changed (optimization) or just overwrite
                # Convert empty string back to None? optional.
                val = cell_val if cell_val != "" else None
                ws.cell(row=r_idx+1, column=c_idx+1, value=val)
                
        wb.save(path)
        return {"status": "ok"}
    except Exception as e:
        return JSONResponse({"status": "error", "msg": str(e)}, status_code=500)



# ==========================================================
# DATA RETRIEVAL (Ported Logic)
# ==========================================================

@router.get("/{project_id}/data/billing")
def get_billing_data(project_id: str, date: str = Query(None), month: str = Query(None)):
    """
    Retrieves billing data for the report, calculated from project files.
    Replaces the old SQLite logic.
    """
    try:
        # 1. Configs
        project_dir = get_project_dir_local(project_id)
        config_device_path = os.path.join(project_dir, "ConfigDevice.json")
        billing_path = os.path.join(project_dir, "config", "billing.json") # Assuming strict struct? Or root?
        # Fallback to root if config/billing.json not found
        if not os.path.exists(billing_path):
             billing_path = os.path.join(project_dir, "billing.json")

        convertors = {}
        if os.path.exists(config_device_path):
             cd = load_json(config_device_path)
             # Transform to old convertors format if needed, or simply use as is.
             # Old format: { "conv_id": { "devices": { "dev_id": { "name": ... } } } }
             # ConfigDevice.json format: { "converters": [ { "id":..., "devices": [...] } ] }
             for c in cd.get("converters", []):
                 cid = str(c.get("id"))
                 c_devices = {}
                 for d in c.get("devices", []):
                     did = str(d.get("id"))
                     c_devices[did] = {"name": d.get("name"), "model": d.get("model")}
                 convertors[cid] = {"devices": c_devices, "name": c.get("name")}

        billing_config = load_json(billing_path, {"price_per_unit": 5.0})
        price = float(billing_config.get("price_per_unit", 5.0))

        # 2. Determine Date
        target_date = date or datetime.now().strftime("%Y-%m-%d")
        target_month = month # YYYY-MM
        
        
        # 3. Read Data (From Billing Database - SQLite)
        # Use the authoritative data from billing_storage to ensure consistency with Billing UI.
        from services.backend.api.billing.billing_storage import get_daily_records, get_monthly_records

        daily_data = {} 
        
        if target_month:
            # Monthly Mode
            # Get aggregated monthly records
            recs = get_monthly_records(project_id, target_month)
            
            # Map to structure
            # For monthly report, we might want daily breakdown or just total?
            # The current frontend expects 'daily' structure for day mode.
            # But the user might be asking for "Month" report which implies 1 row per device for the whole month?
            # Or daily breakdown for that month?
            # Based on frontend: onPeriodValueChange -> if mode=month -> aggregatedDevices = {} (placeholder)
            # It seems frontend support for month view is limited in the JS provided (lines 359-361).
            # But let's return the totals per device for that month in a "summary" object or 
            # if we want to allow day-by-day in month, we'd need get_daily_series logic.
            # Let's stick to returning "devices" map populated with monthly totals for the target_date (which is actually just used as a key).
            # We will use the 'month' string as the key in daily_data to store the month aggregate.
            
            devices_summary = {}
            total_used = 0
            total_cost = 0
            
            for r in recs:
                did = str(r['device_id'])
                used = r['total_energy']
                money = r['total_cost']
                
                total_used += used
                total_cost += money
                
                # Find meta
                dev_name = did
                c_id = "unknown"
                # Lookup name from config
                for cid, cdata in convertors.items():
                    if did in cdata.get("devices", {}):
                        c_id = cid
                        dev_name = cdata["devices"][did].get("name") or dev_name
                        break
                        
                devices_summary[did] = {
                    "device_id": did,
                    "device_name": dev_name,
                    "convertor_id": c_id,
                    "meter_start": r.get('meter_start_month', 0),
                    "meter_end": r.get('meter_end_month', 0),
                    "total_used_today": round(used, 2), # Re-using field name for compatibility
                    "money_today": round(money, 2)
                }
            
            daily_data[target_month] = {
                "devices": devices_summary,
                "total_units_today": round(total_used, 2),
                "total_money_today": round(total_cost, 2)
            }
            
        else:
            # Daily Mode
            recs = get_daily_records(project_id, target_date)
            
            devices_summary = {}
            total_used = 0
            total_cost = 0
            
            for r in recs:
                did = str(r['device_id'])
                used = r['energy_used']
                money = r['cost']
                
                total_used += used
                total_cost += money
                
                # Find meta
                dev_name = did
                c_id = "unknown"
                for cid, cdata in convertors.items():
                    if did in cdata.get("devices", {}):
                        c_id = cid
                        dev_name = cdata["devices"][did].get("name") or dev_name
                        break
                        
                devices_summary[did] = {
                    "device_id": did,
                    "device_name": dev_name,
                    "convertor_id": c_id,
                    "meter_start": r['meter_start'],
                    "meter_end": r['meter_end'],
                    "total_used_today": round(used, 2),
                    "money_today": round(money, 2)
                }
                
            daily_data[target_date] = {
                "devices": devices_summary,
                "total_units_today": round(total_used, 2),
                "total_money_today": round(total_cost, 2)
            }

        # Ensure all configured devices are present (init with 0)
        target_key = target_month if target_month else target_date
        if target_key not in daily_data:
             daily_data[target_key] = { "devices": {}, "total_units_today": 0, "total_money_today": 0 }
             
        target_day_data = daily_data[target_key]
        
        for cid, cdata in convertors.items():
            for did, dinfo in cdata.get("devices", {}).items():
                if did not in target_day_data["devices"]:
                    target_day_data["devices"][did] = {
                        "device_id": did,
                        "device_name": dinfo.get("name") or did,
                        "convertor_id": cid,
                        "meter_start": 0,
                        "meter_end": 0,
                        "total_used_today": 0,
                        "money_today": 0
                    }

        return {
            "status": "ok",
            "data": {
                "daily": daily_data,
                "monthly": {}, # ToDo: Calculate if needed
                "billing": billing_config,
                "convertors": convertors
            }
        }

    except Exception as e:
        import traceback
        traceback.print_exc()
        return JSONResponse({"status": "error", "msg": str(e)}, status_code=500)

# ==========================================================
# CSV GENERATION (Ported Logic)
# ==========================================================

# ==========================================================
# EXCEL GENERATION (Pandas)
# ==========================================================

@router.post("/preview_struct")
def preview_template_struct(
    template_id: str = Body(..., embed=True),
    project_id: str = Body(None, embed=True)
):
    """
    Reads the specified Excel template to determine dynamic columns.
    Scans first 10 rows to find the most likely header row.
    """
    try:
        path = None
        if project_id:
            p_dir = get_templates_dir(project_id)
            p_path = os.path.join(p_dir, template_id)
            if os.path.exists(p_path):
                path = p_path
        
        if not path:
            path = os.path.join(TEMPLATES_DIR, template_id)
            
        if not os.path.exists(path):
            return JSONResponse({"status": "error", "msg": "Template not found"}, status_code=404)

        wb = openpyxl.load_workbook(path, read_only=True)
        ws = wb.active
        
        headers = []
        best_score = 0
        
        # Scan first 20 rows
        keywords = [
            "device", "id", "name", "usage", "unit", "cost", "amount", "price", "meter", 
            "prev", "now", "start", "end", "อุปกรณ์", "มิเตอร์", "หน่วย", "ราคา", "จำนวนเงิน", "ชื่อ",
            "energy", "power", "total", "model", "รุ่น", "เริ่ม", "สิ้นสุด", "ล่าสุด", "ค่าไฟ",
            "รายการ", "รวม", "วันที่", "time", "date", "ลำดับ", "รายละเอียด", "no", "description", "order"
        ]
        
        for row in ws.iter_rows(min_row=1, max_row=20, values_only=True):
            current_row_strs = [str(c).strip() if c is not None else "" for c in row]
            
            # Calculate score: number of non-empty cells that match keywords
            score = 0
            has_content = 0
            is_parameter_row = False
            
            for cell in current_row_strs:
                if not cell: continue
                has_content += 1
                
                # HEAVY PENALTY FOR PLACEHOLDERS
                if "{{" in cell or "}}" in cell:
                    is_parameter_row = True
                    break
                    
                c_lower = cell.lower()
                if any(k in c_lower for k in keywords):
                    score += 1
            
            if is_parameter_row:
                continue
            
            # IGNORE TITLE ROWS:
            # If a row has only 1 non-empty cell, it's likely a title, even if it matches a keyword.
            # Exception: unless the template literally only has 1 column, but unlikely for a report.
            if has_content < 2:
                continue

            if score > best_score:
                best_score = score
                headers = current_row_strs
            # Tie-breaker: prefer row with more content columns if scores equal
            elif score == best_score and score > 0:
                if has_content > len([x for x in headers if x]):
                    headers = current_row_strs
        
        # Fallback if no keywords found: Use row 1 or the first row with > 2 columns
        if not headers:
             for row in ws.iter_rows(min_row=1, max_row=5, values_only=True):
                 r = [str(c).strip() if c else "" for c in row]
                 if len([x for x in r if x]) >= 2:
                     headers = r
                     break

        wb.close()
        
        return {"status": "ok", "headers": headers}

    except Exception as e:
        return JSONResponse({"status": "error", "msg": str(e)}, status_code=500)


@router.get("/generated")
def list_generated_reports(project_id: str = Query(...)):
    """List all generated reports in the project's report directory"""
    try:
        report_dir = get_project_report_dir(project_id)
        if not os.path.exists(report_dir):
            return {"status": "ok", "files": []}
            
        files = []
        for f in os.listdir(report_dir):
            path = os.path.join(report_dir, f)
            if os.path.isfile(path):
                stat = os.stat(path)
                # Determine type from extension
                ext = os.path.splitext(f)[1].lower()
                ftype = "Unknown"
                if ext == ".xlsx": ftype = "Excel"
                elif ext == ".pdf": ftype = "PDF"
                elif ext == ".csv": ftype = "CSV"
                
                files.append({
                    "id": f, # Use filename as ID
                    "name": f,
                    "type": ftype,
                    "date": datetime.fromtimestamp(stat.st_mtime).strftime("%Y-%m-%d %H:%M"),
                    "size": f"{round(stat.st_size / 1024, 1)} KB",
                    "path": f"/api/report/download/{f}?project_id={project_id}" # Virtual path
                })
        
        # Sort by date desc
        files.sort(key=lambda x: x['date'], reverse=True)
        return {"status": "ok", "files": files}
    except Exception as e:
        return JSONResponse({"status": "error", "msg": str(e)}, status_code=500)

@router.get("/download/{filename}")
def download_report(filename: str, project_id: str = Query(...)):
    """Download a specific report file"""
    report_dir = get_project_report_dir(project_id)
    path = os.path.join(report_dir, filename)
    if not os.path.exists(path):
        raise HTTPException(404, "File not found")
    
    return FileResponse(path, filename=filename)

@router.post("/render/excel")
def render_excel_report(
    template_id: str = Body(...), 
    data: dict = Body(...),
    project_id: str = Query(None) 
):
    try:
        # Resolve Template Path
        path = None
        if project_id:
            p_dir = get_templates_dir(project_id)
            p_path = os.path.join(p_dir, template_id)
            if os.path.exists(p_path):
                path = p_path
        
        if not path:
             path = os.path.join(TEMPLATES_DIR, template_id)

        # Output buffer
        output = io.BytesIO()

        if os.path.exists(path):
            # --- TEMPLATE FILLING MODE ---
            wb = openpyxl.load_workbook(path)
            ws = wb.active
            
            # 1. Find Header Row & Start Row
            headers = []
            start_row = 2
            best_score = 0
            best_row_idx = 1
            
            keywords = [
                "device", "id", "name", "usage", "unit", "cost", "amount", "price", "meter", 
                "prev", "now", "start", "end", "อุปกรณ์", "มิเตอร์", "หน่วย", "ราคา", "จำนวนเงิน", "ชื่อ",
                "energy", "power", "total", "model", "รุ่น", "เริ่ม", "สิ้นสุด", "ล่าสุด", "ค่าไฟ",
                "รายการ", "รวม", "วันที่", "time", "date", "ลำดับ", "รายละเอียด", "no", "description", "order"
            ]

            # Scan first 20 rows to find header
            for r_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=20, values_only=True), start=1):
                row_strs = [str(c).strip().lower() if c else "" for c in row]
                score = 0
                has_content = 0
                is_parameter_row = False
                
                for cell in row_strs:
                    if cell: has_content += 1
                    
                    if "{{" in cell or "}}" in cell:
                        is_parameter_row = True
                        break
                        
                    if any(k in cell for k in keywords):
                        score += 1
                
                if is_parameter_row:
                    continue
                
                # IGNORE TITLE ROWS
                if has_content < 2:
                    continue

                if score > best_score:
                    best_score = score
                    headers = row_strs
                    best_row_idx = r_idx
                elif score == best_score and score > 0:
                    if has_content > len([x for x in headers if x]):
                        headers = row_strs
                        best_row_idx = r_idx

            # If we found a header row, data starts after it
            if best_score > 0:
                start_row = best_row_idx + 1
            else:
                # Fallback: Just assume row 1 is header
                start_row = 2
                
            # Prepare flattened device list
            devices_list = []
            if 'devices' in data:
                 devices_list = data['devices']
            elif 'converters' in data:
                 for c in data['converters']:
                     devices_list.extend(c.get('devices', []))
            elif isinstance(data, list): # Raw list
                devices_list = data

            # Write Data
            for i, d in enumerate(devices_list):
                row_idx = start_row + i
                
                # If we have headers, try to map
                if headers:
                    for col_idx, header in enumerate(headers):
                        val = ""
                        # Mapping Logic (Synced with Frontend)
                        if any(k in header for k in ["device", "อุปกรณ์", "ชื่อ", "name", "model", "รุ่น", "รายการ"]):
                            val = d.get("device_name") or d.get("name") or d.get("device_id")
                        elif "id" in header and len(header) < 5: # "ID"
                            val = d.get("device_id")
                        elif any(k in header for k in ["start", "prev", "ก่อน", "เริ่ม"]):
                            val = d.get("meter_prev") or d.get("meter_start")
                        elif any(k in header for k in ["end", "now", "หลัง", "ล่าสุด", "สิ้นสุด"]):
                            val = d.get("meter_now") or d.get("meter_end")
                        elif any(k in header for k in ["used", "usage", "unit", "kwh", "หน่วย", "energy", "พลังงาน", "total"]):
                            val = d.get("used") or d.get("total_used_today")
                        elif any(k in header for k in ["cost", "money", "thb", "baht", "บาท", "ราคา", "ค่าไฟ", "amount"]):
                            val = d.get("money") or d.get("money_today")
                        else:
                            val = d.get(header, "") # Exact match fallback
                        
                        # Only write if val is not empty/None
                        if val is not None and val != "":
                            ws.cell(row=row_idx, column=col_idx+1, value=val)
                else:
                    # No headers found? Default layout
                    ws.cell(row=row_idx, column=1, value=d.get("device_id"))
                    ws.cell(row=row_idx, column=2, value=d.get("device_name"))
                    ws.cell(row=row_idx, column=3, value=d.get("meter_start"))
                    ws.cell(row=row_idx, column=4, value=d.get("meter_end"))
                    ws.cell(row=row_idx, column=5, value=d.get("used"))
                    ws.cell(row=row_idx, column=6, value=d.get("money"))

            wb.save(output)

        else:
            # --- FALLBACK: RAW DUMP (Pandas) ---
            writer = pd.ExcelWriter(output, engine='xlsxwriter')
            
            # Summary Sheet
            summary_data = []
            if 'total_used' in data:
                summary_data.append({"Parameter": "Total Energy (kWh)", "Value": data['total_used']})
            if 'total_money' in data:
                summary_data.append({"Parameter": "Total Cost (THB)", "Value": data['total_money']})
            if summary_data:
                pd.DataFrame(summary_data).to_excel(writer, sheet_name='Summary', index=False)
                
            # Devices Sheet
            devices_list = []
            if 'devices' in data:
                 devices_list = data['devices']
            elif 'converters' in data:
                 for c in data['converters']:
                     devices_list.extend(c.get('devices', []))
            
            if devices_list:
                flat_devs = []
                for d in devices_list:
                    flat_devs.append({
                        "Device ID": d.get("device_id"),
                        "Name": d.get("device_name"),
                        "Meter Start": d.get("meter_prev") or d.get("meter_start"),
                        "Meter End": d.get("meter_now") or d.get("meter_end"),
                        "Used (kWh)": d.get("used"),
                        "Cost (THB)": d.get("money")
                    })
                pd.DataFrame(flat_devs).to_excel(writer, sheet_name='Devices', index=False)
            
            writer.close()

        output.seek(0)
        filename = f"report_{template_id}_{data.get('date', 'export')}.xlsx"
        
        return StreamingResponse(
            output, 
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", 
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )

    except Exception as e:
        import traceback
        traceback.print_exc()
        return JSONResponse({"status": "error", "msg": str(e)}, status_code=500)

@router.post("/render/csv")
def render_csv_report(
    template_id: str = Body(...), 
    data: dict = Body(...)
):
    try:
        output = io.StringIO()
        writer = csv.writer(output)
        output.write('\ufeff')
        
        date_str = data.get('date', 'N/A')
        writer.writerow([f"Report Date: {date_str}"])
        
        devices_list = []
        if 'devices' in data:
             devices_list = data['devices']
        elif 'converters' in data:
             for c in data['converters']:
                 devices_list.extend(c.get('devices', []))
                 
        if devices_list:
            writer.writerow(["Device ID", "Name", "Meter Start", "Meter End", "Used (kWh)", "Cost (THB)"])
            for d in devices_list:
                writer.writerow([
                    d.get("device_id", ""),
                    d.get("device_name", ""),
                    d.get("meter_prev") or d.get("meter_start", 0),
                    d.get("meter_now") or d.get("meter_end", 0),
                    d.get("used", 0),
                    d.get("money", 0)
                ])
        else:
            writer.writerow(["No data found"])
            
        output.seek(0)
        return Response(
            content=output.getvalue(),
            media_type="text/csv",
            headers={"Content-Disposition": f"attachment; filename=report_{date_str}.csv"}
        )
    except Exception as e:
         return JSONResponse({"status": "error", "msg": str(e)}, status_code=500)


# ==========================================================
# HTML RENDERING
# ==========================================================

@router.post("/render")
def render_html_report(
    template_id: str = Body(...), 
    data: dict = Body(...),
    project_id: str = Query(None)
):
    return render_excel_report(template_id=template_id, data=data, project_id=project_id)
